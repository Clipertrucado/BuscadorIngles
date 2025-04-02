from openpyxl import load_workbook
import random

def cargar_datos(archivo_entrada):
    wb = load_workbook(archivo_entrada)
    sheet = wb.active
    encabezados = [cell.value for cell in sheet[1]]
    return wb, sheet, encabezados

def verificar_columnas(encabezados):
    if 'TEMA' not in encabezados:
        raise ValueError("Falta la columna 'TEMA'.")

def filtrar_por_curso(sheet, temas, idx_tema, idx_curso_especifico):
    filas_filtradas = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cumple_tema = not temas or row[idx_tema] in temas  # Si no hay temas, es general
        if cumple_tema and (row[idx_curso_especifico] is None or row[idx_curso_especifico] == ""):
            filas_filtradas.append(row)
    return filas_filtradas

def seleccionar_preguntas(filas_filtradas, num_preguntas):
    preguntas_unicas = []
    vistos = set()

    for fila in filas_filtradas:
        clave = (fila[1], fila[2], fila[6])  # PREGUNTA, OPCIONES, CORRECTA
        if clave not in vistos:
            vistos.add(clave)
            preguntas_unicas.append(fila)

    return preguntas_unicas

def obtener_preguntas(archivo_entrada, temas, num_preguntas, curso_elegido):
    # Si no hay temas, dejamos la lista vacía para filtrar todo
    temas = [int(t.strip()) for t in temas if t.strip().isdigit()]
    wb, sheet, encabezados = cargar_datos(archivo_entrada)

    verificar_columnas(encabezados)
    idx_tema = encabezados.index("TEMA")

    curso_normalizado = curso_elegido.strip().upper().replace('\n', ' ')
    idx_curso = None

    for i, encabezado in enumerate(encabezados):
        if encabezado is None:
            continue
        encabezado_limpio = str(encabezado).strip().upper().replace('\n', ' ')
        if encabezado_limpio == curso_normalizado:
            idx_curso = i
            break

    if idx_curso is None:
        raise ValueError(f"No se encontró la columna '{curso_elegido}' en el archivo.")

    filas_disponibles = filtrar_por_curso(sheet, temas, idx_tema, idx_curso)
    preguntas_unicas = seleccionar_preguntas(filas_disponibles, num_preguntas)

    total_unicas = len(preguntas_unicas)
    if num_preguntas and num_preguntas < total_unicas:
        preguntas_unicas = random.sample(preguntas_unicas, num_preguntas)

    return encabezados, preguntas_unicas, filas_disponibles, total_unicas

def reemplazar_pregunta(filas_disponibles_totales, filas_seleccionadas, indice):
    seleccionadas_set = set(tuple(row) for row in filas_seleccionadas)
    disponibles = [row for row in filas_disponibles_totales if tuple(row) not in seleccionadas_set]
    if not disponibles:
        return None
    nueva = random.choice(disponibles)
    filas_seleccionadas[indice] = nueva
    return nueva

def actualizar_archivo_original(ruta_archivo, filas_actualizadas, encabezados, columna_destino, valor_asignado):
    wb = load_workbook(ruta_archivo)
    sheet = wb.active

    encabezados_normalizados = [str(c).strip().upper().replace('\n', ' ') for c in encabezados]
    col_normalizada = columna_destino.strip().upper().replace('\n', ' ')
    idx_columna = encabezados_normalizados.index(col_normalizada)

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=0):
        contenido_fila = tuple(cell.value for cell in row)
        for seleccionada in filas_actualizadas:
            if contenido_fila[1] == seleccionada[1]:
                row[idx_columna].value = valor_asignado
                break

    wb.save(ruta_archivo)
    return True

def extraer_columnas_curso(encabezados):
    cursos = []
    for col in encabezados:
        if col and isinstance(col, str):
            limpio = col.strip()
            if any(palabra in limpio.upper() for palabra in ["SARGENTO", "OFICIAL", "CABO", "ESCALA", "GC+GJ"]):
                cursos.append(limpio)
    return cursos
